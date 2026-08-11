"""
Excel ingestion service for Service Advisor Performance Tracker.

Supports TWO workbook formats:

1. Original multi-sheet workbook:
   Total Veh, Tyre, Battery, WA, Smiles, IC, EF, Bacta,
   T-Gloss, DIY, ACC, GS to BP Conv, NPS, VOC.

2. New single-sheet workbook:
   Data Sheet

The new Data Sheet contains all advisor/month metrics in one row,
including the product PM Conversion % fields.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from app.services.identity import normalize_name, resolve_identity_key


# ============================================================
# WORKBOOK CONFIGURATION
# ============================================================

SUMMARY_SHEETS = {
    "SA Summary",
    "SA Summary (2)",
}

SINGLE_DATA_SHEET = "Data Sheet"

EXPECTED_SHEETS = [
    "Total Veh",
    "Tyre",
    "Battery",
    "WA",
    "Smiles",
    "IC",
    "EF",
    "Bacta",
    "T-Gloss",
    "DIY",
    "ACC",
    "GS to BP Conv",
    "NPS",
    "VOC",
]

COMMON_COLUMNS = [
    "Month",
    "Branch",
    "City",
    "Employee Number",
    "Name",
    "Designation",
    "Grade",
]


# ============================================================
# OLD MULTI-SHEET FORMAT
# ============================================================

METRIC_COLUMNS = {
    "Total Veh": {
        "GUS": "gus",
        "PM": "pm",
        "10K": "ten_k",
        "20K": "twenty_k",
        "30K": "thirty_k",
    },

    "Tyre": {
        "Tyre": "tyre",
        "Tyre PM Conv %": "tyre_pm_conversion_pct",
    },

    "Battery": {
        "Battery": "battery",
        "Battery PM Conv %": "battery_pm_conversion_pct",
    },

    "WA": {
        "W/A": "wheel_alignment",
        "W/A PM Conv %": "wheel_alignment_pm_conversion_pct",
    },

    "Smiles": {
        "AMC": "smiles_amc",
        "AMC PM Conv %": "smiles_amc_pm_conversion_pct",
    },

    "IC": {
        "IC": "injector_cleaner",
        "IC PM Conv %": "injector_cleaner_pm_conversion_pct",
    },

    "EF": {
        "Engine Flush": "engine_flush",
        "EF PM Conv %": "engine_flush_pm_conversion_pct",
    },

    "Bacta": {
        "Bactakleen": "bactakleen",
        "Bactakleen PM Conv %": "bactakleen_pm_conversion_pct",
        "BK PM Conv %": "bactakleen_pm_conversion_pct",
    },

    "T-Gloss": {
        "VAS Veh (Units)": "vas_vehicles",
        "VAS Treatment (No.s)": "vas_treatments",
        "VAS Value": "vas_value",
        "Body Coating": "body_coating",
        "Ceramic Coating": "ceramic_coating",
        "Graphene": "graphene",
    },

    "DIY": {
        "DIY Product": "diy_value",
    },

    "ACC": {
        "Accessories": "accessories_value",
    },

    "GS to BP Conv": {
        "GS TO BP Conversion Paint Veh.Only": "gs_to_bp_conversion",
    },

    "NPS": {
        "NCDS Sample": "nps_sample",
        "Promoters": "nps_promoters",
        "Neutral": "nps_neutral",
        "Detractor": "nps_detractors",
    },

    "VOC": {
        "IVOC": "ivoc_responses",
        "TKM VOC": "tkm_voc",
    },
}


# ============================================================
# NEW SINGLE-SHEET FORMAT
# ============================================================

SINGLE_SHEET_COLUMNS = {
    "GUS": "gus",
    "PM": "pm",

    "10K": "ten_k",
    "20K": "twenty_k",
    "30K": "thirty_k",

    "Tyre No.s": "tyre",
    "Tyre PM Conv %": "tyre_pm_conversion_pct",

    "Battery No.s": "battery",
    "Battery PM Conv %": "battery_pm_conversion_pct",

    "W/A": "wheel_alignment",
    "W/A PM Conv %": "wheel_alignment_pm_conversion_pct",

    "AMC": "smiles_amc",
    "AMC PM Conv %": "smiles_amc_pm_conversion_pct",

    "IC": "injector_cleaner",
    "IC PM Conv %": "injector_cleaner_pm_conversion_pct",

    "Engine Flush": "engine_flush",
    "EF PM Conv %": "engine_flush_pm_conversion_pct",

    "Bactakleen": "bactakleen",
    "BK PM Conv %": "bactakleen_pm_conversion_pct",

    "VAS Veh (Units)": "vas_vehicles",
    "VAS Treatment (No.s)": "vas_treatments",
    "VAS Value": "vas_value",

    "DIY Product": "diy_value",
    "Accessories": "accessories_value",

    "GS TO BP Conversion Paint Veh.Only": "gs_to_bp_conversion",

    "NCDS Sample": "nps_sample",
    "Promoters": "nps_promoters",
    "Neutral": "nps_neutral",
    "Detractor": "nps_detractors",

    "IVOC": "ivoc_responses",
    "TKM VOC": "tkm_voc",
}


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class ValidationIssue:
    severity: str
    message: str
    sheet: str | None = None
    row: int | None = None


@dataclass
class AdvisorSnapshotPayload:
    month: str
    year: int
    branch: str
    city: str | None
    employee_number: str | None
    display_name: str
    normalized_name: str
    identity_key: str
    designation: str | None
    grade: str | None
    metrics: dict[str, Any] = field(default_factory=dict)


@dataclass
class IngestionResult:
    rows_parsed: int
    snapshots: list[AdvisorSnapshotPayload]
    issues: list[ValidationIssue]


# ============================================================
# VALUE HELPERS
# ============================================================

def _clean_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()

    return text if text else None


def _to_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        raise ValueError("Boolean value is not a valid metric")

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip().replace(",", "")

    if not text:
        return None

    try:
        number = float(text)

        if number.is_integer():
            return int(number)

        return number

    except ValueError as exc:
        raise ValueError(
            f"Invalid numeric value: {value!r}"
        ) from exc


def _to_percentage(value: Any) -> int | float | None:
    """
    Convert Excel percentage values to a dashboard-friendly number.

    Examples:

    75.69  -> 75.69
    "75.69%" -> 75.69
    0.7569 -> 75.69

    Excel percentage cells are sometimes stored internally as
    decimal fractions, so values between 0 and 1 are converted
    to percentage points.
    """

    if value is None or value == "":
        return None

    if isinstance(value, bool):
        raise ValueError("Boolean value is not a valid percentage")

    if isinstance(value, (int, float)):
        number = float(value)

    else:
        text = str(value).strip().replace(",", "")

        if text.endswith("%"):
            text = text[:-1].strip()

        if not text:
            return None

        try:
            number = float(text)

        except ValueError as exc:
            raise ValueError(
                f"Invalid percentage value: {value!r}"
            ) from exc

    # Excel may store 75.69% as 0.7569.
    if 0 < number <= 1:
        number *= 100

    if number.is_integer():
        return int(number)

    return number


# ============================================================
# HEADER HELPERS
# ============================================================

def _header_map(ws) -> dict[str, int]:
    headers = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    return {
        str(value).strip(): index
        for index, value in enumerate(headers)
        if value is not None
    }


def _validate_common_headers(
    headers: dict[str, int],
    sheet_name: str,
) -> list[ValidationIssue]:

    issues: list[ValidationIssue] = []

    for column in COMMON_COLUMNS:
        if column not in headers:
            issues.append(
                ValidationIssue(
                    severity="error",
                    message=f"Missing required column: {column}",
                    sheet=sheet_name,
                )
            )

    return issues


def _validate_metric_headers(
    headers: dict[str, int],
    sheet_name: str,
    metric_columns: dict[str, str],
) -> list[ValidationIssue]:

    issues: list[ValidationIssue] = []

    for source_column in metric_columns:
        if source_column not in headers:
            issues.append(
                ValidationIssue(
                    severity="error",
                    message=f"Missing metric column: {source_column}",
                    sheet=sheet_name,
                )
            )

    return issues


# ============================================================
# ROW PARSER
# ============================================================

def _row_to_record(
    row_values: tuple[Any, ...],
    headers: dict[str, int],
    sheet_name: str,
    row_number: int,
    metric_columns: dict[str, str],
    percentage_columns: set[str] | None = None,
) -> tuple[dict[str, Any] | None, list[ValidationIssue]]:

    issues: list[ValidationIssue] = []

    percentage_columns = percentage_columns or set()

    def value(column: str) -> Any:
        index = headers.get(column)

        if index is None:
            return None

        if index >= len(row_values):
            return None

        return row_values[index]

    month = _clean_text(value("Month"))
    branch = _clean_text(value("Branch"))
    city = _clean_text(value("City"))
    employee_number = _clean_text(value("Employee Number"))
    name = _clean_text(value("Name"))
    designation = _clean_text(value("Designation"))
    grade = _clean_text(value("Grade"))

    if not month:
        issues.append(
            ValidationIssue(
                severity="error",
                message="Missing Month",
                sheet=sheet_name,
                row=row_number,
            )
        )

    if not branch:
        issues.append(
            ValidationIssue(
                severity="error",
                message="Missing Branch",
                sheet=sheet_name,
                row=row_number,
            )
        )

    if not name and not employee_number:
        issues.append(
            ValidationIssue(
                severity="error",
                message="Missing both Employee Number and Name",
                sheet=sheet_name,
                row=row_number,
            )
        )

    if issues:
        return None, issues

    metrics: dict[str, Any] = {}

    for source_column, target_field in metric_columns.items():

        raw_value = value(source_column)

        try:
            if source_column in percentage_columns:
                metrics[target_field] = _to_percentage(raw_value)
            else:
                metrics[target_field] = _to_number(raw_value)

        except ValueError as exc:
            issues.append(
                ValidationIssue(
                    severity="error",
                    message=str(exc),
                    sheet=sheet_name,
                    row=row_number,
                )
            )

    if issues:
        return None, issues

    identity_key = resolve_identity_key(
        employee_number,
        name or "",
    )

    return {
        "month": month,
        "branch": branch,
        "city": city,
        "employee_number": employee_number,
        "display_name": name or employee_number or "",
        "normalized_name": normalize_name(
            name or employee_number or ""
        ),
        "identity_key": identity_key,
        "designation": designation,
        "grade": grade,
        "metrics": metrics,
    }, issues


# ============================================================
# SINGLE-SHEET PARSER
# ============================================================

def _parse_single_sheet(
    workbook,
    year: int,
) -> IngestionResult:

    issues: list[ValidationIssue] = []

    ws = workbook[SINGLE_DATA_SHEET]

    headers = _header_map(ws)

    # Required common columns
    issues.extend(
        _validate_common_headers(
            headers,
            SINGLE_DATA_SHEET,
        )
    )

    # Only require the columns we actually import.
    issues.extend(
        _validate_metric_headers(
            headers,
            SINGLE_DATA_SHEET,
            SINGLE_SHEET_COLUMNS,
        )
    )

    if any(
        issue.severity == "error"
        for issue in issues
    ):
        return IngestionResult(
            rows_parsed=0,
            snapshots=[],
            issues=issues,
        )

    percentage_columns = {
        "Tyre PM Conv %",
        "Battery PM Conv %",
        "W/A PM Conv %",
        "AMC PM Conv %",
        "IC PM Conv %",
        "EF PM Conv %",
        "BK PM Conv %",
    }

    snapshots: dict[str, AdvisorSnapshotPayload] = {}

    for row_number, row_values in enumerate(
        ws.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        if not any(
            value is not None
            for value in row_values
        ):
            continue

        record, row_issues = _row_to_record(
            row_values=row_values,
            headers=headers,
            sheet_name=SINGLE_DATA_SHEET,
            row_number=row_number,
            metric_columns=SINGLE_SHEET_COLUMNS,
            percentage_columns=percentage_columns,
        )

        issues.extend(row_issues)

        if record is None:
            continue

        key = (
            f"{record['month']}|"
            f"{record['identity_key']}"
        )

        if key in snapshots:
            issues.append(
                ValidationIssue(
                    severity="error",
                    message="Duplicate advisor/month in Data Sheet",
                    sheet=SINGLE_DATA_SHEET,
                    row=row_number,
                )
            )
            continue

        snapshots[key] = AdvisorSnapshotPayload(
            month=record["month"],
            year=year,
            branch=record["branch"],
            city=record["city"],
            employee_number=record["employee_number"],
            display_name=record["display_name"],
            normalized_name=record["normalized_name"],
            identity_key=record["identity_key"],
            designation=record["designation"],
            grade=record["grade"],
            metrics=record["metrics"],
        )

    return IngestionResult(
        rows_parsed=len(snapshots),
        snapshots=list(snapshots.values()),
        issues=issues,
    )


# ============================================================
# OLD MULTI-SHEET PARSER
# ============================================================

def _parse_multi_sheet(
    workbook,
    year: int,
) -> IngestionResult:

    issues: list[ValidationIssue] = []

    actual_sheets = set(workbook.sheetnames)

    missing_sheets = [
        sheet
        for sheet in EXPECTED_SHEETS
        if sheet not in actual_sheets
    ]

    for sheet in missing_sheets:
        issues.append(
            ValidationIssue(
                severity="error",
                message=f"Missing required sheet: {sheet}",
                sheet=sheet,
            )
        )

    if missing_sheets:
        return IngestionResult(
            rows_parsed=0,
            snapshots=[],
            issues=issues,
        )

    # --------------------------------------------------------
    # Total Veh = base population
    # --------------------------------------------------------

    base_sheet = workbook["Total Veh"]

    base_headers = _header_map(base_sheet)

    issues.extend(
        _validate_common_headers(
            base_headers,
            "Total Veh",
        )
    )

    issues.extend(
        _validate_metric_headers(
            base_headers,
            "Total Veh",
            METRIC_COLUMNS["Total Veh"],
        )
    )

    if any(
        issue.severity == "error"
        for issue in issues
    ):
        return IngestionResult(
            rows_parsed=0,
            snapshots=[],
            issues=issues,
        )

    snapshots: dict[str, AdvisorSnapshotPayload] = {}

    for row_number, row_values in enumerate(
        base_sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        if not any(
            value is not None
            for value in row_values
        ):
            continue

        record, row_issues = _row_to_record(
            row_values=row_values,
            headers=base_headers,
            sheet_name="Total Veh",
            row_number=row_number,
            metric_columns=METRIC_COLUMNS["Total Veh"],
        )

        issues.extend(row_issues)

        if record is None:
            continue

        key = (
            f"{record['month']}|"
            f"{record['identity_key']}"
        )

        if key in snapshots:
            issues.append(
                ValidationIssue(
                    severity="error",
                    message="Duplicate advisor/month in Total Veh",
                    sheet="Total Veh",
                    row=row_number,
                )
            )
            continue

        snapshots[key] = AdvisorSnapshotPayload(
            month=record["month"],
            year=year,
            branch=record["branch"],
            city=record["city"],
            employee_number=record["employee_number"],
            display_name=record["display_name"],
            normalized_name=record["normalized_name"],
            identity_key=record["identity_key"],
            designation=record["designation"],
            grade=record["grade"],
            metrics=record["metrics"],
        )

    # --------------------------------------------------------
    # Merge remaining sheets
    # --------------------------------------------------------

    for sheet_name in EXPECTED_SHEETS:

        if sheet_name == "Total Veh":
            continue

        ws = workbook[sheet_name]

        sheet_headers = _header_map(ws)

        sheet_issues = []

        sheet_issues.extend(
            _validate_common_headers(
                sheet_headers,
                sheet_name,
            )
        )

        sheet_issues.extend(
            _validate_metric_headers(
                sheet_headers,
                sheet_name,
                METRIC_COLUMNS[sheet_name],
            )
        )

        issues.extend(sheet_issues)

        if any(
            issue.severity == "error"
            for issue in sheet_issues
        ):
            continue

        percentage_columns = {
            column
            for column in METRIC_COLUMNS[sheet_name]
            if "Conv %" in column
        }

        for row_number, row_values in enumerate(
            ws.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):

            if not any(
                value is not None
                for value in row_values
            ):
                continue

            record, row_issues = _row_to_record(
                row_values=row_values,
                headers=sheet_headers,
                sheet_name=sheet_name,
                row_number=row_number,
                metric_columns=METRIC_COLUMNS[sheet_name],
                percentage_columns=percentage_columns,
            )

            issues.extend(row_issues)

            if record is None:
                continue

            key = (
                f"{record['month']}|"
                f"{record['identity_key']}"
            )

            snapshot = snapshots.get(key)

            if snapshot is None:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        message=(
                            "Advisor/month exists in metric "
                            "sheet but not Total Veh"
                        ),
                        sheet=sheet_name,
                        row=row_number,
                    )
                )
                continue

            snapshot.metrics.update(
                record["metrics"]
            )

    return IngestionResult(
        rows_parsed=len(snapshots),
        snapshots=list(snapshots.values()),
        issues=issues,
    )


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def parse_workbook(
    file_path: str | Path,
    year: int = 2026,
) -> IngestionResult:

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Excel file not found: {path}"
        )

    workbook = None

    try:
        workbook = openpyxl.load_workbook(
            path,
            data_only=True,
            read_only=True,
        )

        actual_sheets = set(workbook.sheetnames)

        # ====================================================
        # NEW FORMAT
        # ====================================================

        if SINGLE_DATA_SHEET in actual_sheets:

            return _parse_single_sheet(
                workbook,
                year,
            )

        # ====================================================
        # OLD FORMAT
        # ====================================================

        return _parse_multi_sheet(
            workbook,
            year,
        )

    finally:

        # Important on Windows:
        # close workbook before temporary upload file
        # is removed.
        if workbook is not None:
            workbook.close()